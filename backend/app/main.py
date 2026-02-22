from __future__ import annotations

import json
import logging
from contextlib import asynccontextmanager
from io import BytesIO
from typing import Any

import uvicorn
from fastapi import FastAPI, File, Form, HTTPException, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse

from app.config import DEFAULT_DOMAIN_DENYLIST, get_env_config, get_runtime_port
from app.database import Database
from app.job_manager import JobManager
from app.schemas import (
    AppConfigResponse,
    CacheClearResponse,
    JobCreateResponse,
    JobSettings,
    JobStatusResponse,
    ReviewResponse,
    ReviewSubmitRequest,
)
from app.services.ct_registry import CTRegistryService
from app.services.domain_lookup import DomainLookupService
from app.services.http_client import RetryHttpClient
from app.services.social_hint import SocialHintService

if not logging.getLogger().handlers:
    logging.basicConfig(level=logging.INFO)

logger = logging.getLogger("startdate_finder.main")


@asynccontextmanager
async def lifespan(app: FastAPI):
    env = app.state.env
    db = Database(env["db_path"])
    http_client = RetryHttpClient()
    manager = JobManager(
        db=db,
        ct_registry=CTRegistryService(
            db,
            http_client,
            soda_app_token=env["soda_app_token"],
            test_mode=env["enable_test_mode"],
        ),
        domain_lookup=DomainLookupService(
            db,
            http_client,
            whoisxml_api_key=env["whoisxml_api_key"],
            test_mode=env["enable_test_mode"],
        ),
        social_hints=SocialHintService(http_client, test_mode=env["enable_test_mode"]),
    )
    app.state.env = env
    app.state.manager = manager
    app.state.http_client = http_client

    yield

    await manager.close()
    await http_client.close()


app = FastAPI(title="StartDate Finder API", version="1.0.0", lifespan=lifespan)

app_env = get_env_config()
app.state.env = app_env

# Initial production posture for easier first deploy:
# keep wildcard CORS enabled, then tighten to your specific frontend origin.
# Recommended origin to lock down later:
# https://YOUR_USERNAME.github.io
cors_setting = app_env["cors_allow_origins"].strip()
if cors_setting == "*" or not cors_setting:
    allow_origins = ["*"]
else:
    allow_origins = [item.strip() for item in cors_setting.split(",") if item.strip()]

app.add_middleware(
    CORSMiddleware,
    allow_origins=allow_origins,
    allow_credentials=False,
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.get("/api/health")
async def health() -> dict[str, str]:
    return {"status": "ok"}


@app.get("/api/config", response_model=AppConfigResponse)
async def config() -> AppConfigResponse:
    env_cfg = app.state.env
    return AppConfigResponse(
        defaults=JobSettings(
            high_confidence_threshold=0.85,
            prefer_earliest_known_date=False,
            enable_rdap_lookup=True,
            enable_whois_fallback=bool(env_cfg["whoisxml_api_key"]),
            enable_social_hints=False,
            min_plausible_date=env_cfg["min_plausible_date"],
            denylist_domains=list(DEFAULT_DOMAIN_DENYLIST),
        ),
        whois_key_present=bool(env_cfg["whoisxml_api_key"]),
        feature_social_hints_env=bool(env_cfg["feature_social_hints"]),
    )


@app.post("/api/jobs", response_model=JobCreateResponse)
async def create_job(
    file: UploadFile = File(...),
    settings_json: str = Form(default="{}"),
) -> JobCreateResponse:
    filename = (file.filename or "").lower()
    if not filename.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Only .xlsx uploads are supported")

    try:
        incoming = json.loads(settings_json or "{}")
    except json.JSONDecodeError as exc:
        raise HTTPException(status_code=400, detail=f"Invalid settings JSON: {exc}") from exc

    try:
        settings = JobSettings.model_validate(incoming)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Invalid settings: {exc}") from exc

    if settings.enable_social_hints and not app.state.env["feature_social_hints"]:
        settings.enable_social_hints = False

    file_bytes = await file.read()
    if not file_bytes:
        raise HTTPException(status_code=400, detail="Uploaded file is empty")

    try:
        # quick format sanity check
        from openpyxl import load_workbook

        load_workbook(BytesIO(file_bytes), read_only=True)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Invalid XLSX file: {exc}") from exc

    job_id = app.state.manager.create_job(file_bytes, settings)
    return JobCreateResponse(job_id=job_id)


@app.get("/api/jobs/{job_id}/status", response_model=JobStatusResponse)
async def job_status(job_id: str) -> JobStatusResponse:
    payload = app.state.manager.get_status(job_id)
    if not payload:
        raise HTTPException(status_code=404, detail="Job not found")
    return JobStatusResponse.model_validate(payload)


@app.get("/api/jobs/{job_id}/events")
async def job_events(job_id: str) -> StreamingResponse:
    status = app.state.manager.get_status(job_id)
    if not status:
        raise HTTPException(status_code=404, detail="Job not found")

    async def generator():
        async for event in app.state.manager.stream_events(job_id):
            yield event

    return StreamingResponse(generator(), media_type="text/event-stream")


@app.get("/api/jobs/{job_id}/review", response_model=ReviewResponse)
async def review(job_id: str) -> ReviewResponse:
    status = app.state.manager.get_status(job_id)
    if not status:
        raise HTTPException(status_code=404, detail="Job not found")
    rows = app.state.manager.get_review_rows(job_id)
    return ReviewResponse(job_id=job_id, rows=rows)


@app.post("/api/jobs/{job_id}/review")
async def submit_review(job_id: str, request: ReviewSubmitRequest) -> dict[str, Any]:
    status = app.state.manager.get_status(job_id)
    if not status:
        raise HTTPException(status_code=404, detail="Job not found")
    await app.state.manager.submit_review(
        job_id,
        selections=[selection.model_dump() for selection in request.selections],
    )
    updated = app.state.manager.get_status(job_id)
    return {"ok": True, "status": updated}


@app.get("/api/jobs/{job_id}/download")
async def download(job_id: str) -> StreamingResponse:
    status = app.state.manager.get_status(job_id)
    if not status:
        raise HTTPException(status_code=404, detail="Job not found")
    if status["status"] not in {"completed"}:
        raise HTTPException(status_code=400, detail="Job is not ready for download")

    try:
        content = await app.state.manager.download_workbook(job_id)
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Unable to generate download: {exc}") from exc

    headers = {"Content-Disposition": f'attachment; filename="startdate-finder-{job_id}.xlsx"'}
    return StreamingResponse(
        BytesIO(content),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )


@app.post("/api/cache/clear", response_model=CacheClearResponse)
async def clear_cache() -> CacheClearResponse:
    await app.state.manager.clear_cache()
    return CacheClearResponse(cleared=True)


if __name__ == "__main__":
    port = get_runtime_port(default=8000)
    logger.info("Starting StartDate Finder API on 0.0.0.0:%s", port)
    uvicorn.run("app.main:app", host="0.0.0.0", port=port)
