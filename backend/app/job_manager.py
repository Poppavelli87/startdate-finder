from __future__ import annotations

import asyncio
import json
import uuid
from collections.abc import AsyncGenerator
from io import BytesIO
from typing import Any

from openpyxl import load_workbook

from app.database import Database
from app.schemas import JobCounts, JobSettings
from app.services.ct_registry import CTRegistryService
from app.services.domain_lookup import DomainLookupService
from app.services.selection import choose_start_date
from app.services.social_hint import SocialHintService
from app.utils import normalize_business_name

AUDIT_COLUMNS = [
    "chosen_start_date",
    "chosen_source",
    "confidence",
    "ct_matched_name",
    "ct_entity_id",
    "ct_registration_date_raw",
    "ct_query_notes",
    "domain",
    "domain_created_date",
    "domain_lookup_notes",
    "social_profile_url",
    "social_created_hint_date",
    "social_lookup_notes",
]


class JobManager:
    def __init__(
        self,
        db: Database,
        ct_registry: CTRegistryService,
        domain_lookup: DomainLookupService,
        social_hints: SocialHintService,
    ) -> None:
        self._db = db
        self._ct_registry = ct_registry
        self._domain_lookup = domain_lookup
        self._social_hints = social_hints
        self._tasks: dict[str, asyncio.Task[None]] = {}
        self._events: dict[str, asyncio.Queue[dict[str, Any]]] = {}
        self._source_files: dict[str, bytes] = {}

    def create_job(self, file_bytes: bytes, settings: JobSettings) -> str:
        job_id = str(uuid.uuid4())
        self._db.create_job(job_id, settings.model_dump())
        self._events[job_id] = asyncio.Queue()
        self._source_files[job_id] = file_bytes
        task = asyncio.create_task(self._run_job(job_id, settings))
        self._tasks[job_id] = task
        self._publish_status(job_id)
        return job_id

    def get_status(self, job_id: str) -> dict[str, Any] | None:
        job = self._db.get_job(job_id)
        if not job:
            return None
        progress_total = max(0, int(job["progress_total"]))
        progress_done = max(0, int(job["progress_done"]))
        progress_pct = round((progress_done / progress_total) * 100.0, 1) if progress_total else 0.0
        return {
            "job_id": job["job_id"],
            "status": job["status"],
            "progress_done": progress_done,
            "progress_total": progress_total,
            "progress_pct": progress_pct,
            "message": job.get("message", ""),
            "counts": job.get("counts", _default_counts()),
            "can_download": job["status"] == "completed",
            "error": job.get("error"),
        }

    def get_review_rows(self, job_id: str) -> list[dict[str, Any]]:
        return self._db.list_review_rows(job_id)

    async def submit_review(self, job_id: str, selections: list[dict[str, Any]]) -> None:
        job = self._db.get_job(job_id)
        if not job:
            raise ValueError("Job not found")
        settings = JobSettings.model_validate(job.get("settings", {}))
        rows = self._db.list_job_rows(job_id)
        row_map = {row["row_index"]: row for row in rows}

        for selection in selections:
            row_index = int(selection["row_index"])
            entry = row_map.get(row_index)
            if not entry:
                continue
            result = dict(entry["result"])
            ct = dict(result.get("ct", {}))
            candidates = entry.get("candidates") or result.get("ct_candidates", [])
            no_match = bool(selection.get("no_match"))
            candidate_index = selection.get("candidate_index")
            if not no_match and candidate_index is not None and 0 <= int(candidate_index) < len(candidates):
                chosen = candidates[int(candidate_index)]
                ct.update(
                    {
                        "matched_name": chosen.get("name", ""),
                        "entity_id": chosen.get("entity_id", ""),
                        "registration_date_raw": chosen.get("registration_date_raw", ""),
                        "registration_date": chosen.get("registration_date", ""),
                        "confidence": float(chosen.get("confidence") or 0.0),
                        "similarity": float(chosen.get("similarity") or 0.0),
                    }
                )
                ct["query_notes"] = _append_note(
                    ct.get("query_notes", ""),
                    "ct_review_selection_applied",
                )
                result["needs_review"] = False
            else:
                ct["query_notes"] = _append_note(
                    ct.get("query_notes", ""),
                    "ct_review_marked_no_match",
                )
                result["needs_review"] = False

            selection_result = choose_start_date(
                ct,
                result.get("domain_lookup", {}),
                result.get("social_lookup", {}),
                settings,
            )
            result["ct"] = ct
            result["audit"] = _build_audit(result, ct, result.get("domain_lookup", {}), result.get("social_lookup", {}), selection_result)
            result["date_established"] = result["audit"]["chosen_start_date"]

            self._db.update_job_row_result(
                job_id=job_id,
                row_index=row_index,
                result=result,
                needs_review=False,
            )

        updated_rows = self._db.list_job_rows(job_id)
        counts = _compute_counts(updated_rows)
        self._db.update_job(job_id, counts=counts, message="Review updates applied")
        self._publish_status(job_id)

    async def download_workbook(self, job_id: str) -> bytes:
        source = self._source_files.get(job_id)
        if source is None:
            raise ValueError("Job file no longer available in memory")
        workbook = load_workbook(BytesIO(source))
        sheet = workbook.active
        headers = _read_headers(sheet)
        header_map = _ensure_headers(sheet, headers)
        date_col_idx = header_map.get("Date Established")

        rows = self._db.list_job_rows(job_id)
        row_map = {row["row_index"]: row for row in rows}
        for row_index, payload in row_map.items():
            result = payload["result"]
            audit = result.get("audit", {})
            if date_col_idx:
                sheet.cell(row=row_index, column=date_col_idx, value=audit.get("chosen_start_date", ""))
            for column in AUDIT_COLUMNS:
                col_idx = header_map[column]
                sheet.cell(row=row_index, column=col_idx, value=audit.get(column, ""))

        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        return output.getvalue()

    async def clear_cache(self) -> None:
        self._db.clear_cache()

    async def stream_events(self, job_id: str) -> AsyncGenerator[str, None]:
        queue = self._events.get(job_id)
        if queue is None:
            yield _sse_payload({"error": "job_not_found"})
            return

        self._publish_status(job_id)
        while True:
            try:
                event = await asyncio.wait_for(queue.get(), timeout=15.0)
            except TimeoutError:
                status = self.get_status(job_id)
                if not status:
                    break
                yield _sse_payload(status)
                if status["status"] in {"completed", "failed"}:
                    break
                continue

            yield _sse_payload(event)
            if event.get("status") in {"completed", "failed"}:
                break

    async def close(self) -> None:
        tasks = [task for task in self._tasks.values() if not task.done()]
        for task in tasks:
            task.cancel()
        if tasks:
            await asyncio.gather(*tasks, return_exceptions=True)

    async def _run_job(self, job_id: str, settings: JobSettings) -> None:
        try:
            source = self._source_files[job_id]
            workbook = load_workbook(BytesIO(source), data_only=True)
            sheet = workbook.active
            headers = _read_headers(sheet)
            header_map = _normalize_header_map(headers)

            if "Business" not in header_map:
                raise ValueError("Input sheet must include a 'Business' column")

            rows = _extract_rows(sheet, headers)
            total_rows = len(rows)
            running_counts = _default_counts()
            running_counts["total_rows"] = total_rows
            self._db.update_job(
                job_id,
                status="running",
                progress_total=total_rows,
                progress_done=0,
                counts=running_counts,
                message="Processing spreadsheet",
            )
            self._publish_status(job_id)

            ct_cache: dict[str, dict[str, Any]] = {}
            domain_cache: dict[str, dict[str, Any]] = {}
            processed = 0

            for source_row in rows:
                result = await self._enrich_row(source_row, settings, ct_cache, domain_cache)
                self._db.upsert_job_row(
                    job_id=job_id,
                    row_index=int(source_row["_row_index"]),
                    source_row=source_row,
                    result=result,
                    needs_review=bool(result.get("needs_review")),
                    candidates=result.get("ct_candidates", []),
                )
                processed += 1
                _increment_counts(running_counts, result)
                self._db.update_job(
                    job_id,
                    progress_done=processed,
                    counts=running_counts,
                    message=f"Processed {processed}/{total_rows}",
                )
                self._publish_status(job_id)

            self._db.update_job(
                job_id,
                status="completed",
                message="Processing completed",
                counts=running_counts,
            )
            self._publish_status(job_id)
        except Exception as exc:
            self._db.update_job(
                job_id,
                status="failed",
                message="Processing failed",
                error=f"{exc.__class__.__name__}: {exc}",
            )
            self._publish_status(job_id)

    async def _enrich_row(
        self,
        source_row: dict[str, Any],
        settings: JobSettings,
        ct_cache: dict[str, dict[str, Any]],
        domain_cache: dict[str, dict[str, Any]],
    ) -> dict[str, Any]:
        business = _as_text(source_row.get("Business"))
        city = _as_text(source_row.get("City"))
        zip_code = _as_text(source_row.get("Zip"))
        url = _as_text(source_row.get("URL"))
        social_url_column = _as_text(source_row.get("Social URL"))

        ct_key = f"{normalize_business_name(business)}|{city.strip().upper()}|{zip_code.strip()}"
        ct_result = _empty_ct_result()
        if business.strip():
            if ct_key in ct_cache:
                ct_result = ct_cache[ct_key]
            else:
                ct_result = await self._ct_registry.lookup(business, city, zip_code)
                ct_cache[ct_key] = ct_result

        should_try_domain = (
            not ct_result.get("registration_date")
            or float(ct_result.get("confidence") or 0.0) < settings.high_confidence_threshold
        )

        domain_result = _empty_domain_result()
        if should_try_domain and (settings.enable_rdap_lookup or settings.enable_whois_fallback):
            domain_key = url.strip().lower()
            if domain_key and domain_key in domain_cache:
                domain_result = domain_cache[domain_key]
            else:
                domain_result = await self._domain_lookup.lookup_from_url(
                    url,
                    denylist_domains=settings.denylist_domains,
                    enable_rdap_lookup=settings.enable_rdap_lookup,
                    enable_whois_fallback=settings.enable_whois_fallback,
                )
                if domain_key:
                    domain_cache[domain_key] = domain_result
        else:
            domain_result = {
                "domain": "",
                "domain_created_date": "",
                "source": "",
                "lookup_notes": "domain_lookup_not_needed",
            }

        social_result = _empty_social_result()
        if settings.enable_social_hints:
            social_url = social_url_column or _infer_social_url(url)
            if social_url:
                social_result = await self._social_hints.lookup(social_url)
            else:
                social_result = {
                    "social_profile_url": "",
                    "social_created_hint_date": "",
                    "social_lookup_notes": "social_hint_not_attempted",
                    "confidence": 0.0,
                }

        selection = choose_start_date(ct_result, domain_result, social_result, settings)
        audit = _build_audit(source_row, ct_result, domain_result, social_result, selection)
        return {
            "date_established": audit["chosen_start_date"],
            "audit": audit,
            "needs_review": bool(ct_result.get("needs_review")),
            "ct_candidates": ct_result.get("candidates", []),
            "ct": ct_result,
            "domain_lookup": domain_result,
            "social_lookup": social_result,
        }

    def _publish_status(self, job_id: str) -> None:
        queue = self._events.get(job_id)
        status = self.get_status(job_id)
        if queue is None or status is None:
            return
        try:
            queue.put_nowait(status)
        except asyncio.QueueFull:
            pass


def _build_audit(
    source_row: dict[str, Any],
    ct_result: dict[str, Any],
    domain_result: dict[str, Any],
    social_result: dict[str, Any],
    selection: dict[str, Any],
) -> dict[str, Any]:
    return {
        "chosen_start_date": selection.get("chosen_start_date", ""),
        "chosen_source": selection.get("chosen_source", "not_found"),
        "confidence": float(selection.get("confidence", 0.0)),
        "ct_matched_name": ct_result.get("matched_name", ""),
        "ct_entity_id": ct_result.get("entity_id", ""),
        "ct_registration_date_raw": ct_result.get("registration_date_raw", ""),
        "ct_query_notes": selection.get("ct_query_notes", ct_result.get("query_notes", "")),
        "domain": domain_result.get("domain", ""),
        "domain_created_date": domain_result.get("domain_created_date", ""),
        "domain_lookup_notes": selection.get("domain_lookup_notes", domain_result.get("lookup_notes", "")),
        "social_profile_url": social_result.get("social_profile_url", ""),
        "social_created_hint_date": social_result.get("social_created_hint_date", ""),
        "social_lookup_notes": selection.get("social_lookup_notes", social_result.get("social_lookup_notes", "")),
    }


def _compute_counts(rows: list[dict[str, Any]]) -> dict[str, int]:
    counts = _default_counts()
    counts["total_rows"] = len(rows)
    for row in rows:
        _increment_counts(counts, row["result"])
    return counts


def _increment_counts(counts: dict[str, int], result: dict[str, Any]) -> None:
    audit = result.get("audit", {})
    chosen_source = audit.get("chosen_source", "not_found")
    needs_review = bool(result.get("needs_review"))

    if chosen_source == "ct_registry" and not needs_review:
        counts["auto_matched"] += 1
    if needs_review:
        counts["needs_review"] += 1
    if chosen_source == "not_found":
        counts["not_found"] += 1
    if chosen_source in {"domain_rdap", "whoisxml"}:
        counts["filled_via_domain"] += 1
    if chosen_source == "social_hint":
        counts["filled_via_social"] += 1


def _default_counts() -> dict[str, int]:
    return JobCounts().model_dump()


def _read_headers(sheet: Any) -> list[str]:
    headers: list[str] = []
    for cell in sheet[1]:
        value = "" if cell.value is None else str(cell.value).strip()
        headers.append(value)
    return headers


def _normalize_header_map(headers: list[str]) -> dict[str, int]:
    return {header: idx for idx, header in enumerate(headers) if header}


def _ensure_headers(sheet: Any, headers: list[str]) -> dict[str, int]:
    final_headers = list(headers)
    if "Date Established" not in final_headers:
        final_headers.append("Date Established")
    for col in AUDIT_COLUMNS:
        if col not in final_headers:
            final_headers.append(col)
    for idx, header in enumerate(final_headers, start=1):
        sheet.cell(row=1, column=idx, value=header)
    return {header: idx for idx, header in enumerate(final_headers, start=1)}


def _extract_rows(sheet: Any, headers: list[str]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for row_index in range(2, sheet.max_row + 1):
        values = [sheet.cell(row=row_index, column=col_index).value for col_index in range(1, len(headers) + 1)]
        if not any(value is not None and str(value).strip() for value in values):
            continue
        payload: dict[str, Any] = {headers[idx]: values[idx] for idx in range(len(headers)) if headers[idx]}
        payload["_row_index"] = row_index
        rows.append(payload)
    return rows


def _as_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _append_note(existing: str, note: str) -> str:
    if not existing:
        return note
    return f"{existing}; {note}"


def _infer_social_url(url: str) -> str:
    lower = (url or "").lower()
    if any(host in lower for host in ["facebook.com/", "instagram.com/", "x.com/", "twitter.com/", "linkedin.com/company/"]):
        return url
    return ""


def _empty_ct_result() -> dict[str, Any]:
    return {
        "matched_name": "",
        "entity_id": "",
        "registration_date_raw": "",
        "registration_date": "",
        "confidence": 0.0,
        "similarity": 0.0,
        "query_notes": "",
        "needs_review": False,
        "candidates": [],
    }


def _empty_domain_result() -> dict[str, Any]:
    return {
        "domain": "",
        "domain_created_date": "",
        "source": "",
        "lookup_notes": "",
    }


def _empty_social_result() -> dict[str, Any]:
    return {
        "social_profile_url": "",
        "social_created_hint_date": "",
        "social_lookup_notes": "",
        "confidence": 0.0,
    }


def _sse_payload(data: dict[str, Any]) -> str:
    return f"data: {json.dumps(data)}\n\n"

